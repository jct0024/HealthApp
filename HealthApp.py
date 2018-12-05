# -*- coding: utf-8 -*-
import openpyxl as op;
import numpy as np;
#Cargamos la base de datos
doc = op.load_workbook("BaseDeDatosDeAlimentos.xlsx");
print(doc.sheetnames)
#Seleccionamos la hoja de excell que contendrá dicha información-
hoja1 = doc['Hoja1'];
#Inicializamos la matriz y las filas
print(hoja1.max_row)
print(hoja1.max_column)
nFila=-1;
mAlimentos = np.empty((hoja1.max_row,hoja1.max_column+1), dtype=object);
for alimentos in hoja1:
    nFila =nFila+1
    if (nFila!=0):
        i = 0;
        for inf in alimentos:
            if (i==0):
                mAlimentos[nFila,i]=inf.value;
            elif (i==1):
                mAlimentos[nFila,i]=inf.value;
            elif (i==2):
                mAlimentos[nFila,i]=inf.value;
            elif (i==3):
                mAlimentos[nFila,i]=inf.value;
            elif (i==4):
                mAlimentos[nFila,i]=inf.value;
            elif (i==5):
                mAlimentos[nFila,i]=inf.value;
            elif (i==6):
                mAlimentos[nFila,i]=inf.value;
            elif (i==7):
                mAlimentos[nFila,i]=inf.value;
            i=i+1
        #Iniciamos a cero la frecuencia con la que ingieren los alimentos.
        #Esta varaible luego servirá para imlementar un LRU, para la gestión de alimentos.
        mAlimentos[nFila,8] = 0;
            
print(nFila)
print(i)
    
print (mAlimentos)